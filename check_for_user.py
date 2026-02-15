import openpyxl as op

def check(name,pas):
    filename="users.xlsx"
    wb=op.load_workbook(filename)
    ws=wb["Info"]

    for i in range(2,100000):
        if ws[f"A{i}"].value==None:
            break
        if ws[f"A{i}"].value==name:
            if ws[f"C{i}"].value==int(pas):
                return ws[f"B{i}"].value
    return ""