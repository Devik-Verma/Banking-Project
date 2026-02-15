#This segment is only used to fetch the mail of user whose currently depositing or withdrawing money 

import openpyxl as op

filename="users.xlsx"

def gettingmail(n):
    name=n
    wb=op.load_workbook(filename)
    ws=wb["Info"]
    for i in range(2,100000):
        if ws[f"A{i}"]==None:
            break
        if ws[f"A{i}"].value==name:
            return ws[f"D{i}"].value
        
