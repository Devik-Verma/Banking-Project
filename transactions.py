#import the required modules
import openpyxl as op
import tkinter as tk 
import datetime as dt
import mailing
import fetch_the_mail
import otp

def sendotp():
    lblfortop.config(text="Please enter the OTP")
    ent3.config(state="normal")
    otp.generate(ent2.get())
    
#pushing entries to database
def dep():
    date=dt.datetime.now().strftime("%Y-%m-%d")
 
    #disable the entry widgets
    ent1.config(state="readonly")
    ent2.config(state="readonly")
    ent3.config(state="readonly")
    
    filename="users.xlsx"

    wb=op.load_workbook(filename)
    ws=wb["Info"]
    ws2=wb["Deposit"]

    #otp from the txt file
    with open("containotp.txt","r") as file:
        otpfromtxt = file.readline().strip()

    for i in range(2,1000000):
        #breaks the loop as soon as an empty row is found to avoid large iterations
        if ws[f"A{i}"].value==None:
            break
        if name==ws[f"A{i}"].value:
            if str(ent3.get())==str(otpfromtxt):
                ws[f"B{i}"].value+=int(ent1.get())
                ws2.append([date,int(ent1.get()),name])
                lbl3.config(text=f"Amount of {ent1.get()} has been successfully deposited")

                wb.save(filename)
                #mailing the user
                mail=fetch_the_mail.gettingmail(name)
                mailing.sendmail1(name,mail,ent1.get())
            else:
                lbl3.config(text="Incorrect OTP entered")
    

#pushing entries to database
def wit():
    date=dt.datetime.now().strftime("%Y-%m-%d")

    ent1.config(state="readonly")
    ent2.config(state="readonly")
    ent3.config(state="readonly")

    filename="users.xlsx"

    wb=op.load_workbook(filename)
    ws=wb["Info"]
    ws2=wb["Withdraw"]

    #otp from the txt file
    with open("containotp.txt","r") as file:
        otpfromtxt = file.readline().strip()

    for i in range(2,1000000):
        #breaks the loop as soon as an empty row is found to avoid large iterations
        if ws[f"A{i}"].value==None:
            break
        if name==ws[f"A{i}"].value:
            if str(ent3.get())==str(otpfromtxt):
                #if the balance is low 
                if int(ent1.get())>ws[f"B{i}"].value:
                    lbl3.config(text="Insufficient balance")
                else:
                    ws[f"B{i}"].value-=int(ent1.get())
                    ws2.append([date,int(ent1.get()),name])
                    lbl3.config(text=f"Amount of {ent1.get()} has been successfully withdrawn")

                    #mailing the user
                    mail=fetch_the_mail.gettingmail(name)
                    mailing.sendmail2(name,mail,ent1.get())
            else:
                lbl3.config(text="Incorrect OTP entered")
    
    wb.save(filename)


#tkinter segment
def deposit(n):
    global name
    global ent1
    global ent2
    global lbl3
    global ent3
    global lblfortop

    name=n

    page5=tk.Tk()
    
    frame1=tk.Frame(page5)#grid
    frame2=tk.Frame(page5)#pack
    
    lbl1=tk.Label(frame1,text="Enter amount to be deposited")
    lbl1.grid(row=0,column=0)

    ent1=tk.Entry(frame1)
    ent1.grid(row=0,column=1)

    lbl2=tk.Label(frame1,text="Please enter your phone number")#phone number of the user
    lbl2.grid(row=1,column=0)

    ent2=tk.Entry(frame1)
    ent2.grid(row=1,column=1)

    getotp=tk.Button(frame1,text="Get OTP",command=sendotp)
    getotp.grid(row=1,column=2)

    lblfortop=tk.Label(frame1,text="")#otp from the user
    lblfortop.grid(row=2,column=0)

    ent3=tk.Entry(frame1,state="readonly")
    ent3.grid(row=2,column=1)

    btn1=tk.Button(frame2,text="Proceed",command=dep)
    btn1.pack()

    lbl3=tk.Label(frame2,text="")
    lbl3.pack()
    
    frame1.pack(pady=10)
    frame2.pack(pady=10)

    page5.mainloop()

#tkinter segment
def withdraw(n):
    global name
    global ent1
    global ent2
    global lbl3


    name=n

    page5=tk.Tk()
    
    frame1=tk.Frame(page5)#grid
    frame2=tk.Frame(page5)#pack
    
    lbl1=tk.Label(frame1,text="Enter amount to be withdrawn")
    lbl1.grid(row=0,column=0)

    ent1=tk.Entry(frame1)
    ent1.grid(row=0,column=1)

    lbl2=tk.Label(frame1,text="Please enter your phone number")
    lbl2.grid(row=1,column=0)

    ent2=tk.Entry(frame1)
    ent2.grid(row=1,column=1)

    getotp=tk.Button(frame1,text="Get OTP",command=sendotp)
    getotp.grid(row=1,column=2)

    lblforotp=tk.Label(frame1,text="")
    lblforotp.grid(row=1,column=0)

    ent3=tk.Entry(frame1)
    ent3.grid(row=1,column=1)

    btn1=tk.Button(frame2,text="Proceed",command=wit)
    btn1.pack()

    lbl3=tk.Label(frame2,text="")
    lbl3.pack()
    
    frame1.pack(pady=10)
    frame2.pack(pady=10)

    page5.mainloop()
